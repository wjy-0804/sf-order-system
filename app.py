#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
顺丰快递模版录入系统 - Flask 后端
"""

import json
import re
import os
import io
import copy
import csv
import tempfile
import glob
import sqlite3
import functools
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file, session, redirect, url_for
from flask import Response, render_template_string
import openpyxl
import xlrd  # 支持读取旧版 .xls 文件
from openpyxl import load_workbook

app = Flask(__name__, static_folder='.', static_url_path='')
app.secret_key = 'sf-order-system-2026-secret-key'

# ======================== 数据库初始化 ========================
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'orders.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    # 用户表
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            is_admin INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
    ''')
    # 兼容旧库升级：如果 is_admin 列不存在则添加
    try:
        c.execute('SELECT is_admin FROM users LIMIT 1')
    except sqlite3.OperationalError:
        c.execute('ALTER TABLE users ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0')
    # 订单表：status='pending' 未导出，'exported' 已导出
    c.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            order_data TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            batch_id TEXT,
            created_at TEXT NOT NULL,
            exported_at TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    # 地址簿分组表
    c.execute('''
        CREATE TABLE IF NOT EXISTS address_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    # 地址簿明细表
    c.execute('''
        CREATE TABLE IF NOT EXISTS address_book (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            group_id INTEGER NOT NULL,
            name TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            address TEXT NOT NULL DEFAULT '',
            product TEXT NOT NULL DEFAULT '',
            quantity TEXT NOT NULL DEFAULT '1',
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (group_id) REFERENCES address_groups(id) ON DELETE CASCADE
        )
    ''')
    # 创建默认管理员账号 lmy123 / lmy123（超级管理员）
    c.execute('SELECT id, is_admin FROM users WHERE username = ?', ('lmy123',))
    existing = c.fetchone()
    if not existing:
        c.execute(
            'INSERT INTO users (username, password, is_admin, created_at) VALUES (?, ?, 1, ?)',
            ('lmy123', 'lmy123', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        )
    elif not existing['is_admin']:
        c.execute('UPDATE users SET is_admin = 1 WHERE username = ?', ('lmy123',))
    conn.commit()
    conn.close()

init_db()

# ======================== 登录验证装饰器 ========================
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': '请先登录'}), 401
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

# ======================== 管理员验证装饰器 ========================
def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '请先登录'}), 401
        if not session.get('is_admin'):
            return jsonify({'success': False, 'error': '需要管理员权限'}), 403
        return f(*args, **kwargs)
    return decorated

SF_TEMPLATE_PATH = os.environ.get(
    'SF_TEMPLATE_PATH',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), '顺丰模版.xlsx')
)
STATS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'stats')
os.makedirs(STATS_DIR, exist_ok=True)

# 顺丰模版列头（按顺序）
SF_COLUMNS = [
    '用户订单号', '寄件公司', '寄件人', '寄件电话', '寄件详细地址',
    '收件公司', '收件人', '收件电话', '收件详细地址',
    '托寄物内容1', '托寄物数量1', '托寄物单价1', '托寄物编码1',
    '托寄物内容2', '托寄物数量2', '托寄物单价2', '托寄物编码2',
    '寄方备注', '月结卡号', '运费付款方式', '业务类型', '件数', '包裹重量',
    '件数1', '包裹重量1', '长（cm）1', '宽（cm）1', '高（cm）1',
    '件数2', '包裹重量2', '长（cm）2', '宽（cm）2', '高（cm）2',
    '代收卡号', '代收金额', '保价类型', '保价金额', '包装服务',
    '签回单-纸质回单', '签回单-拍照回传', '自取件', '拍照回传',
    '是否超长超重', '超长超重服务费', '是否大件入户', '大件入户服务费',
    '保鲜服务', '保单配送', '票据专送', '密钥认证', '密钥认证类型',
    '身份证后6位', '双人派送', '等通知派送', '是否定时派送', '派送日期',
    '派送时段', '温度追溯', '珍宝服务', '到付现结优惠', '到付现结卡号',
    '委托类型', '委托人姓名', '委托人电话', '委托人地址', '打包服务',
    '溯源服务', '安装服务', '安装服务类型', '精温服务', '宅配延伸',
    '宅配延伸月结卡号', '宅配延伸服务编码', '准时保',
    '扩展字段1', '扩展字段2', '扩展字段3', '扩展字段4', '扩展字段5',
    '其他费用', '预约揽件日期', '预约揽件时间', '预约揽件收派员工号',
    '标准模板完整版标记勿删'
]

def parse_orders(raw_text):
    """
    智能解析订单文本，支持：
    1. 纯文本格式
    2. 简单表格（Tab/空格分隔）
    3. 复杂表格（CSV/Excel复制内容）
    返回订单列表，每个订单是dict
    """
    # 统一换行符：浏览器textarea发送 \r\n，必须标准化为 \n
    # 否则 _split_order_blocks_v2 的 \n{2,} 无法匹配 \r\n\r\n
    raw_text = raw_text.replace('\r\n', '\n').replace('\r', '\n')

    orders = []
    
    # 尝试检测是否是表格格式（含Tab分隔符）
    lines = raw_text.strip().split('\n')
    lines = [l.strip() for l in lines if l.strip()]
    
    if not lines:
        return orders
    
    # 检测分隔符
    tab_count = sum(1 for l in lines[:5] if '\t' in l)
    comma_count = sum(1 for l in lines[:5] if ',' in l)
    
    if tab_count >= 2:
        orders = parse_table(lines, '\t')
    elif comma_count >= 2:
        orders = parse_table(lines, ',')
    else:
        orders = parse_freetext(raw_text)
    
    return orders


def parse_table(lines, sep):
    """解析表格格式订单"""
    orders = []
    if not lines:
        return orders
    
    # 第一行判断是否是表头
    header_line = lines[0].split(sep)
    header_map = {}
    
    # 核心识别字段（订单号/收件公司/备注等非必要字段不在此列，需用户手动配置规则时再加）
    # 别名顺序：精确表头名放最前面（优先score=3精确匹配），通用短别名放后面
    field_aliases = {
        '收件人': [
            '收货人/提货人姓名',  # 微店导出标准（精确！含/分隔符）
            '收件人姓名',  # 有赞/电商标准
            '收货人姓名', '收件人', '收货人',  # 通用
            '联系人', '姓名', '买家姓名', '客户姓名', 'name', 'receiver', '客户',
        ],
        '收件电话': [
            '收件人电话',  # 花城/帮会等导出标准（精确！区别于"寄件人电话"）
            '收货人/提货人手机号',  # 微店导出标准（精确！含/分隔符）
            '收货人手机号码', '收件人手机号码', '收货人手机号',  # 电商标准（精确！含"收货人×"前缀）
            '收货人电话',  # 乡伴/有赞等电商导出标准
            '手机号码',
            '联系电话', '收件电话', '收货电话', '手机号',
            '手机', '电话', 'phone', 'mobile', 'tel', '联系方式',
        ],
        '收件详细地址': [
            '收件人全地址',  # 花城/帮会等导出标准（精确！完整的省市区+街道，区别于"收件人街道地址"）
            '收货/提货详细地址',  # 微店导出标准（精确！含/分隔符）
            '收货人完整地址', '收件人完整地址', '收货人详细地址',  # 电商标准（含"收货人×"/"收件人×"前缀）
            '收货人地址',  # 乡伴/有赞等电商导出标准
            '收件详细地址', '收货详细地址',
            '收货地址', '地址', '详细地址',
            'address', '收件地址', '送货地址', '送达地址',
        ],
        '托寄物内容1': [
            '产品',  # 花城/帮会等导出标准（精确！区别于"规格"列）
            'SKU规格', 'SKU名称',  # 电商标准（精确！）
            '商品规格', '商品名称', '规格',  # 带/不带前缀的规格列
            '托寄物内容1', '货物', '物品', '品名', '商品',
            'product', 'goods', 'item', '内容', '货品', '寄件内容',
        ],
        '托寄物数量1': [
            '商品件数',  # 微店/电商标准（精确！区别于"商品总件数"）
            'SKU数量',  # 电商标准（精确！）
            '托寄物数量1', '数量', '件数', 'qty', 'quantity', '重量', '数',
        ],
    }
    
    # Bug18修复: 商品+规格双列合并 — _spec_col 需在 if/else 之前初始化
    _spec_col = None

    # 尝试识别表头（双向匹配：别名∈单元格 或 单元格∈别名）
    is_header = False
    for cell in header_line:
        cell_clean = cell.strip().lower()
        if not cell_clean:
            continue
        for field, aliases in field_aliases.items():
            if any(a.lower() == cell_clean
                   or a.lower() in cell_clean  # 别名是单元格的子串（如"收件人" in "收件人姓名"）
                   or cell_clean in a.lower()   # 单元格是别名的子串
                   for a in aliases):
                is_header = True
                break
        if is_header:
            break
    
    if is_header:
        # 构建列索引映射（全局最优匹配 + 别名精确度排序）
        # Bug14修复：对每个表头cell，遍历所有field找得分最高的
        # Bug15修复：同分时按命中别名的**长度**排序，越长的别名越精确
        #   例如 "商品名称"(4字) > "规格"(2字)，"收件人手机号码"(6字) > "手机"(2字)
        header_map = {}  # col_idx -> field_name
        _match_score = {}  # field_name -> (col_idx, score, alias_len) 用于去重和精度排序
        for i, cell in enumerate(header_line):
            cell_clean = cell.strip()
            if not cell_clean:
                continue  # 跳过空单元格，避免 '' in alias 恒True 的bug

            # 遍历所有field，找该cell的最佳匹配（记录命中别名的长度）
            best_field = None
            best_score = 0
            best_alias_len = 0  # 命中别名的字符长度（用于同分时的二级排序）

            # 寄件/发件人列不能匹配收件字段（防止"寄件人电话"被误匹配为收件电话）
            _is_sender_col = ('寄件' in cell_clean or '发件' in cell_clean)
            _receiver_fields = {'收件人', '收件电话', '收件详细地址', '收件公司'}

            for field, aliases in field_aliases.items():
                # 寄件人列跳过收件字段
                if _is_sender_col and field in _receiver_fields:
                    continue

                is_exact = any(a.lower() == cell_clean.lower() for a in aliases)
                is_sub_in_cell = any(a.lower() in cell_clean.lower() for a in aliases)  # 别名是单元格子串
                is_cell_in_alias = any(cell_clean.lower() in a.lower() for a in aliases)  # 单元格是别名子串
                if is_exact or is_sub_in_cell or is_cell_in_alias:
                    score = 3 if is_exact else (2 if is_sub_in_cell else 1)
                    # 找到当前命中的别名长度
                    hit_len = 0
                    if is_exact:
                        hit_len = max(len(a) for a in aliases if a.lower() == cell_clean.lower())
                    elif is_sub_in_cell:
                        hit_len = max(len(a) for a in aliases if a.lower() in cell_clean.lower())
                    elif is_cell_in_alias:
                        hit_len = len(cell_clean)

                    # 排序：score > alias_len（分数优先，同分时别名越长越精确）
                    if score > best_score or (score == best_score and hit_len > best_alias_len):
                        best_score = score
                        best_field = field
                        best_alias_len = hit_len

            if best_field:
                existing = _match_score.get(best_field)
                # 替换条件：分数更高，或 同分且别名更长（更精确）
                # 同分同长度时保留先匹配到的列（不再被后面的覆盖）
                should_replace = False
                if existing is None:
                    should_replace = True
                elif best_score > existing[1]:
                    should_replace = True
                elif best_score == existing[1] and best_alias_len > existing[2]:
                    should_replace = True

                if should_replace:
                    header_map[i] = best_field
                    _match_score[best_field] = (i, best_score, best_alias_len)
                    if existing and existing[0] in header_map and header_map[existing[0]] == best_field:
                        del header_map[existing[0]]

        # === SKU优先规则: 当表格同时有 SKU 和 商品名称 列时，SKU 优先作为托寄物内容1 ===
        # 德风等电商导出常见模式：SKU 列是精简商品名（如"有机黄甜玉米 10斤"），
        # 商品名称 列是完整描述（如"广东有机黄甜玉米 4.5斤｜现代生态种植..."）
        # 此时应取 SKU 而非商品名称，也不应合并两列
        if '托寄物内容1' in _match_score:
            _prod_col = _match_score['托寄物内容1'][0]
            _prod_col_text = header_line[_prod_col].strip() if _prod_col < len(header_line) else ''
            # 检查是否存在独立 SKU 列
            _sku_col = None
            for i, cell in enumerate(header_line):
                if not cell or i == _prod_col:
                    continue
                if cell.strip().upper() == 'SKU':
                    _sku_col = i
                    break
            # 当前映射的是商品名称列，且存在 SKU 列 → 切换到 SKU
            if _sku_col is not None and ('商品名称' in _prod_col_text or '商品名' in _prod_col_text):
                del header_map[_prod_col]
                header_map[_sku_col] = '托寄物内容1'
                _match_score['托寄物内容1'] = (_sku_col, 1, 3)

        # Bug18修复: 商品+规格双列合并 — 某些电商表格将商品名和规格分两列（如 col=商品名称, col+1=规格）
        # 检测是否有"规格"类型列被托寄物内容1匹配到但未入选（被"商品名称"等更长别名覆盖）
        if '托寄物内容1' in _match_score:
            prod_main_col = _match_score['托寄物内容1'][0]
            _spec_keywords = ('规格', 'SKU规格', 'spec', '型号', 'size')
            for i, cell_clean in enumerate(header_line):
                if i == prod_main_col or not cell_clean or i in header_map:
                    continue
                cell_lower = cell_clean.lower()
                is_spec_type = any(
                    sk.lower() == cell_lower or sk.lower() in cell_lower or cell_lower in sk.lower()
                    for sk in _spec_keywords
                )
                if is_spec_type:
                    _spec_col = i
                    break

        data_lines = lines[1:]
    else:
        data_lines = lines
    
    for line in data_lines:
        if not line.strip():
            continue
        cells = line.split(sep)
        order = {}
        
        if header_map:
            for i, cell in enumerate(cells):
                if i in header_map:
                    order[header_map[i]] = cell.strip()
        else:
            # 无表头时，按位置猜测
            # 尝试从每个cell中提取信息
            full_text = ' '.join(cells)
            order = extract_from_text(full_text)
        
        # 补充提取缺失字段
        if not order.get('收件人') or not order.get('收件电话') or not order.get('收件详细地址'):
            full_text = sep.join(cells)
            extracted = extract_from_text(full_text)
            for k, v in extracted.items():
                if not order.get(k):
                    order[k] = v
        
        # Bug18修复: 商品名 + 规格自动合并（如"有机甜玉米" + "4.5斤" → "有机甜玉米4.5斤"）
        if _spec_col is not None and _spec_col < len(cells) and '托寄物内容1' in order:
            spec_val = str(cells[_spec_col]).strip() if cells[_spec_col] else ''
            if spec_val and order['托寄物内容1']:
                order['托寄物内容1'] = order['托寄物内容1'].strip() + spec_val

        # 数据值清洗（去掉常见的前缀/后缀噪音）
        if '托寄物内容1' in order and order['托寄物内容1']:
            val = order['托寄物内容1']
            # 去掉 "规格:" / "规格：" 前缀（电商SKU常见格式）
            for prefix in ('规格:', '规格：', '【', '['):
                if val.startswith(prefix):
                    val = val[len(prefix):]
            # 去掉末尾分号/句号等
            val = val.rstrip(';；,，。.')
            order['托寄物内容1'] = val.strip()

        # Bug18修复: 检测是否缺少规格信息（商品名中无数字/重量单位时提醒用户手动补充）
        if order.get('托寄物内容1'):
            import re as _re
            has_weight_info = bool(_re.search(r'[\d.]+\s*(斤|克|kg|g|斤装|kg装|个|件|箱|包|袋|盒)',
                                            order['托寄物内容1']))
            if not has_weight_info:
                order['_needs_spec_hint'] = True

        if any(order.values()):
            orders.append(order)
    
    return orders


def extract_from_text(text):
    """从自由文本中提取关键字段"""
    order = {}
    
    # 提取手机号
    phone_pattern = r'1[3-9]\d{9}|0\d{2,3}[-\s]\d{7,8}|\d{11}'
    phones = re.findall(phone_pattern, text)
    if phones:
        order['收件电话'] = phones[0]
    
    # 提取省市区地址
    addr_pattern = r'[^\s，,。；;]+?(?:省|自治区)[^\s，,。；;]*?(?:市|州)[^\s，,。；;]*?(?:区|县|市)[^\s，,。；;]{0,50}'
    addr_match = re.search(addr_pattern, text)
    if addr_match:
        order['收件详细地址'] = addr_match.group()
    else:
        # 尝试更宽松的地址匹配
        addr_pattern2 = r'(?:地址|收货地址|送货地址)[：:]\s*([^\n，,]{5,60})'
        addr_match2 = re.search(addr_pattern2, text)
        if addr_match2:
            order['收件详细地址'] = addr_match2.group(1).strip()
    
    # 提取姓名（在手机号附近的短文本，2-4个汉字）
    name_patterns = [
        r'(?:收件人|收货人|姓名|联系人)[：:\s]*([^\s，,。0-9]{2,6})',
        r'([^\s，,。0-9]{2,4})(?:\s+|，|,)?' + (phones[0] if phones else r'\d{11}') if phones else r'',
    ]
    for pattern in name_patterns:
        if not pattern:
            continue
        name_match = re.search(pattern, text)
        if name_match:
            order['收件人'] = name_match.group(1).strip()
            break
    
    if not order.get('收件人'):
        # 寻找短汉字序列
        chinese_names = re.findall(r'[\u4e00-\u9fa5]{2,4}', text)
        if chinese_names:
            # 排除常见地名关键词
            exclude = ['省', '市', '区', '县', '路', '街', '号', '楼', '镇', '乡', '村', '收件', '寄件', '地址', '商品', '货物', '手机', '电话']
            for name in chinese_names:
                if not any(ex in name for ex in exclude):
                    order['收件人'] = name
                    break
    
    # 提取商品名称
    goods_patterns = [
        r'(?:商品|货物|物品|品名|托寄物)[：:\s]*([^\n，,。]{2,20})',
        r'(?:寄|发)[：:\s]*([^\n，,。]{2,20})',
    ]
    for pattern in goods_patterns:
        goods_match = re.search(pattern, text)
        if goods_match:
            order['托寄物内容1'] = goods_match.group(1).strip()
            break
    
    # 提取订单号
    order_patterns = [
        r'(?:订单号|单号|order)[：:\s#]*([A-Za-z0-9\-]{6,20})',
        r'NO[.:：]?\s*([A-Za-z0-9\-]{6,20})',
    ]
    for pattern in order_patterns:
        order_match = re.search(pattern, text, re.IGNORECASE)
        if order_match:
            order['用户订单号'] = order_match.group(1).strip()
            break
    
    return order


def parse_freetext(text):
    """解析自由文本格式订单 — 多策略级联解析引擎 v3

    支持的格式：
    A. 结构化键值对：收件人:/手机号码:/详细地址:  (必须含"收件人"或"手机号码")
    B. "地址："前缀：地址：xxx，姓名电话 \\n 商品
    C. "姓名 电话 地址"：姓名在前，空格/粘连分隔
    D. "地址 姓名 电话"：地址在前，同行或跨行
    E. 兜底模糊提取
    """
    orders = []
    raw_blocks = _split_order_blocks_v2(text)

    for block in raw_blocks:
        block = block.strip()
        if not block or len(block) < 5:
            continue

        # 先检测是否包含重复订单（如两个相同的结构化块粘在一起）
        sub_orders = _try_split_duplicate_structured(block)
        if len(sub_orders) > 1:
            for sub in sub_orders:
                order = _dispatch_and_parse(sub.strip())
                if order and any(order.values()):
                    orders.append(order)
            continue

        order = _dispatch_and_parse(block)
        if order and any(order.values()):
            orders.append(order)

    return orders


def _strip_receiver_label(block):
    """
    剥离块首的「收货人信息:」/「收件人信息:」等标签行前缀。
    支持两种形式：
      1. 标签与姓名在同一行：「收货人信息: 张小雪 138xxxx」→ 「张小雪 138xxxx」
      2. 标签独占一行：「收货人信息:\n张小雪 138xxxx」→ 「张小雪 138xxxx」
    """
    # 匹配行首的 "收货人信息:"/"收件人信息:"/"收货人:"/"收件人信息:" 等标签
    label_pattern = re.compile(
        r'^(?:收货人信息|收件人信息|收货信息|收件信息)\s*[：:]\s*',
        re.MULTILINE
    )
    lines = block.split('\n')
    first_line = lines[0].strip()
    m = label_pattern.match(first_line)
    if m:
        remainder = first_line[m.end():].strip()
        if remainder:
            # 标签和内容在同一行，直接把标签去掉
            lines[0] = remainder
        else:
            # 标签独占一行，删掉这行
            lines = lines[1:]
        block = '\n'.join(lines)
    return block


def _normalize_phone_spaces(text):
    """规范化手机号内部空格：'189 2913 2348' → '18929132348'

    只处理明显的手机号格式（1[3-9]开头+数字/空格混合），不影响其他文本。
    """
    def _fix_phone(m):
        raw = m.group(0)
        digits = re.sub(r'\s+', '', raw)
        if len(digits) == 11:
            return digits
        return raw
    # 匹配1[3-9]开头、后续为数字和空格混合的11位左右片段
    return re.sub(r'1[3-9][\d ]{8,12}', _fix_phone, text)


def _dispatch_and_parse(block):
    """判断block格式类型，分发给对应解析器"""
    # 预处理：剥离「收货人信息:」等标签前缀（不影响其他格式）
    block = _strip_receiver_label(block)
    # 预处理：规范化手机号内空格（"189 2913 2348" → "18929132348"）
    block = _normalize_phone_spaces(block)
    fmt = _classify_block_format(block)

    if fmt == 'structured':
        order = _parse_structured(block)
    elif fmt == 'addr_prefix':
        order = _parse_addr_prefix(block)
    elif fmt == 'name_phone_addr':
        order = _parse_name_phone_addr(block)
    elif fmt == 'addr_name_phone':
        order = _parse_addr_name_phone(block)
    elif fmt == 'phone_name':
        order = _parse_phone_name(block)
    elif fmt == 'contact_phone':
        order = _parse_contact_phone(block)
    else:
        order = _parse_fallback(block)

    # 所有路径统一提取商品（如果还没提取到）
    if order:
        _extract_product_from_other_lines(block, order)

    return order


def _classify_block_format(block):
    """分类block的格式类型"""
    lines = [l.strip() for l in block.split('\n') if l.strip()]
    if not lines:
        return 'unknown'

    first_line = lines[0]

    # === 优先检测 "姓名在前" 的格式（Type C）===
    # 因为地址行也可能包含姓名+电话，必须先排除姓名在首的情况
    # --- 类型C："姓名 电话 地址"（姓名在行首）---
    # C1: "姓名 电话 地址" 空格分隔
    if re.match(r'^[\u4e00-\u9fa5]{2,4}\s+1[3-9]\d{9}\s+', first_line):
        return 'name_phone_addr'
    # C2: "姓名 电话(粘连) 地址" 或 "姓名 电话(无尾空格)地址"
    if re.match(r'^[\u4e00-\u9fa5]{2,4}\s*1[3-9]\d{9}', first_line):
        return 'name_phone_addr'
    if (re.match(r'^[\u4e00-\u9fa5]{2,4}\s*1[3-9]\d{9}\s*$', first_line)
        and len(lines) >= 2):
        return 'name_phone_addr'

    # --- 类型A：结构化键值对 ---
    # 收件人关键词必须在行首（强信号），其他键可以出现在行中间（支持 / 分隔格式）
    has_receiver_kw = bool(re.search(r'^(?:收件人|收货人|姓名|联系人)[：:\s]', block, re.MULTILINE))
    has_phone_kw = bool(re.search(r'(?:手机号码|收件电话|收货电话|手机|电话号码|联系电话|联系方式|手机号)[：:\s]', block))
    has_addr_detail_kw = bool(re.search(r'(?:详细地址|所在地区|省市区)[：:\s]', block))
    if has_receiver_kw and (has_phone_kw or has_addr_detail_kw):
        return 'structured'

    # --- 类型B："地址："前缀 ---
    if re.match(r'^\s*(?:地址|收件地址|收货地址|送货地址|详细地址)[：:\s]', first_line):
        return 'addr_prefix'

    # --- 新增：电话+姓名粘连格式（首行 = 电话+姓名，无分隔）---
    # 例："13822153541廖琼"  "13800138000张三"
    if re.match(r'^1[3-9]\d{9}[一-龥]{2,4}$', first_line):
        return 'phone_name'

    # --- 新增：联系电话前缀格式 ---
    # 格式：地址 + 联系电话：phone + 姓名 + 商品
    # 例："广东省 东莞市 寮步镇...   联系电话：18929132348   李生 蔬菜包6斤"
    if re.search(r'(?:联系电话|联系方式|收件电话|收货电话)[：:\s]+\d', first_line):
        return 'contact_phone'

    # === 最后检测 "地址在前" 的格式（Type D）===
    # --- 类型D："地址 姓名 电话" ---
    if (re.search(r'(?:省|自治区|市).{0,20}(?:市|区|县|镇|乡)', first_line)
        and re.search(r'[\u4e00-\u9fa5]{2,4}\s*1[3-9]\d{9}\s*$', first_line)):
        return 'addr_name_phone'
    if (len(lines) >= 2
        and re.search(r'(?:省|自治区|市).{0,15}(?:市|区|县|镇)', lines[0])
        and (re.search(r'1[3-9]\d{9}', lines[1]) or re.match(r'^[\u4e00-\u9fa5]{2,4}', lines[1]))):
        return 'addr_name_phone'
    if (re.search(r'(?:省|自治区|市).{10,}(?:市|区|县|镇)', first_line)
        and re.search(r'[\u4e00-\u9fa5]{2,4}1[3-9]\d{9}\s*$', first_line)):
        return 'addr_name_phone'

    return 'unknown'


def _split_order_blocks_v2(text):
    """将文本智能拆分为多个订单block"""
    blocks = re.split(r'\n{2,}', text)
    if len(blocks) <= 3:
        sub_blocks = []
        for b in blocks:
            # 用前瞻确保"数字."后面跟的是空格或中文（避免把"4.5斤"中的4.当成列表编号）
            parts = re.split(r'(?m)^(?:\d+)[.、。](?=\s|[\u4e00-\u9fa5])', b)
            for p in parts:
                p = p.strip()
                if p:
                    sub_blocks.append(p)
        if len(sub_blocks) > len(blocks):
            blocks = sub_blocks
    return blocks


def _try_split_duplicate_structured(block):
    """检测一个block内是否包含多个重复的结构化订单（如两个刘琴）"""
    receivers = list(re.finditer(r'^(?:收件人|收货人|姓名)[：:\s]', block, re.MULTILINE))
    if len(receivers) >= 2:
        positions = [m.start() for m in receivers]
        result = []
        for i, pos in enumerate(positions):
            end = positions[i + 1] if i + 1 < len(positions) else len(block)
            chunk = block[pos:end].strip()
            if chunk:
                result.append(chunk)
        if len(result) > 1:
            return result
    return [block]


# ========================================================================
# 格式A: 结构化键值对
# ========================================================================

def _split_kv_line(line):
    """将一行可能包含多对"key:value / key2:value2"的文本拆分为多个(key, value)元组

    支持 / 和空格 分隔的多键值对
    """
    line = line.strip()
    if not line:
        return []

    # 检测是否包含 / 分隔符且看起来像多键值对（value部分还有 key:value 模式）
    if '/' in line:
        # 按 / 分割，然后对每段提取 key: value
        parts = [p.strip() for p in line.split('/')]
        results = []
        for part in parts:
            m = re.match(r'^([^：:\s]{2,10})[：:\s]+(.{1,100})$', part.strip())
            if m:
                results.append((m.group(1).strip(), m.group(2).strip()))
        if len(results) >= 2:
            return results

    # 单键值对的常规处理
    m = re.match(r'^([^：:\s]{2,10})[：:\s]+(.{1,100})$', line)
    if m:
        return [(m.group(1).strip(), m.group(2).strip())]

    return []


def _parse_structured(block):
    """解析结构化键值对格式"""
    lines = block.split('\n')
    order = {}
    region_val = ''

    structured_keys = {
        '收件人': ['收件人', '收货人', '姓名', '联系人', '客户姓名'],
        '收件电话': ['收件电话', '收货电话', '手机号码', '手机', '电话号码',
                     '联系电话', '联系方式', '手机号'],
        '收件详细地址': ['详细地址', '收件地址', '收货地址', '送货地址'],
        '所在地区': ['所在地区', '地区', '省市区', '所属区域'],
        '收件公司': ['收件公司', '收货公司', '公司名称', '公司', '单位'],
        '托寄物内容1': ['商品名称', '商品', '货物', '物品', '品名', '内容',
                        '寄件物', '托寄物', '产品'],
        '用户订单号': ['订单号', '单号', '订单编号', '编号'],
        '寄方备注': ['备注', '备注信息', '说明', '特殊说明'],
    }

    for line in lines:
        line = line.strip()
        if not line:
            continue
        kv_pairs = _split_kv_line(line)
        for key_raw, val in kv_pairs:
            matched_field = None
            for field, aliases in structured_keys.items():
                if key_raw in aliases or key_raw == field:
                    matched_field = field
                    break
            if not matched_field:
                continue

            if matched_field == '所在地区':
                region_val = val
            elif matched_field == '收件详细地址':
                if region_val and not val.startswith(region_val):
                    order[matched_field] = region_val + val
                else:
                    order[matched_field] = val
            else:
                order[matched_field] = val

    return order


# ========================================================================
# 格式B: "地址："前缀
# ========================================================================

def _parse_addr_prefix(block):
    """处理"地址：xxx，姓名电话"格式"""
    addr_line_match = re.search(
        r'(?:地址|收件地址|收货地址|送货地址|详细地址)[：:\s]+(.+?)(?:\n|$)',
        block
    )
    if not addr_line_match:
        return {}

    order = {}
    addr_content = addr_line_match.group(1).strip()
    phone_in_addr = re.search(r'(1[3-9]\d{9})\s*$', addr_content)

    if phone_in_addr:
        order['收件电话'] = phone_in_addr.group(1)
        remaining = addr_content[:phone_in_addr.start()].rstrip(' ，,；;、')
        name_match = re.search(r'([\u4e00-\u9fa5]{2,4})\s*$', remaining)
        exclude_kw = _get_address_exclude_kw()
        if name_match:
            candidate_name = name_match.group(1)
            if not any(kw in candidate_name for kw in exclude_kw):
                order['收件人'] = candidate_name
                addr_text = remaining[:name_match.start()].rstrip(' ，,；;、')
            else:
                addr_text = remaining
        else:
            addr_text = remaining
        order['收件详细地址'] = _strip_trailing_product(addr_text).strip()
    else:
        order['收件详细地址'] = _strip_trailing_product(addr_content)
        phones = re.findall(r'1[3-9]\d{9}', addr_content)
        if phones:
            order['收件电话'] = phones[0]
        _extract_name_smart(order, block)

    return order


# ========================================================================
# 格式C: "姓名 电话 地址"
# ========================================================================

def _parse_name_phone_addr(block):
    """处理"姓名 电话 地址"格式"""
    lines = [l.strip() for l in block.split('\n') if l.strip()]
    if not lines:
        return {}

    order = {}
    first_line = lines[0]

    # C1: "姓名 电话 地址" 空格分隔 同行
    m = re.match(r'^([\u4e00-\u9fa5]{2,4})\s+(1[3-9]\d{9})\s+(.{5,120})$', first_line)
    if m:
        order['收件人'] = m.group(1)
        order['收件电话'] = m.group(2)
        addr_raw = m.group(3).strip()
        stripped = _strip_trailing_product(addr_raw)
        order['收件详细地址'] = stripped
        # 如果有被剥离的商品部分，保存为商品
        if len(stripped) < len(addr_raw):
            product_part = addr_raw[len(stripped):].strip()
            if product_part:
                order['托寄物内容1'] = product_part
        return order

    # C2: "姓名 电话(粘连/半空格) 地址" — 允许姓名与电话间有0-N个空格
    m = re.match(r'^([\u4e00-\u9fa5]{2,4})\s*(1[3-9]\d{9})(.{5,120})$', first_line)
    if m:
        order['收件人'] = m.group(1)
        order['收件电话'] = m.group(2)
        addr_raw = m.group(3).strip()
        stripped = _strip_trailing_product(addr_raw)
        order['收件详细地址'] = stripped
        # 如果有被剥离的商品部分，保存为商品
        if len(stripped) < len(addr_raw):
            product_part = addr_raw[len(stripped):].strip()
            if product_part:
                order['托寄物内容1'] = product_part
        return order

    # C3: 跨行 - 第一行"姓名 电话"，后续行为地址
    m = re.match(r'^([\u4e00-\u9fa5]{2,4})\s*(1[3-9]\d{9})\s*$', first_line)
    if m and len(lines) >= 2:
        order['收件人'] = m.group(1)
        order['收件电话'] = m.group(2)
        addr_parts = []
        for l in lines[1:]:
            l = l.strip()
            if not l:
                continue
            # 如果这行同时像地址又像商品（地址末尾带商品描述）
            if _looks_like_product(l) and re.search(r'(?:省|自治区|市).{0,20}(?:市|区|县|镇)', l):
                # 剥离商品后作为地址，精确提取被剥离的商品名
                stripped = _strip_trailing_product(l)
                addr_parts.append(stripped)
                # 提取被剥离的商品部分：原文本去除地址后剩余的内容
                product_part = l[len(stripped):].strip() if len(stripped) < len(l) else ''
                if product_part:
                    order.setdefault('托寄物内容1', product_part)
                else:
                    # 兜底：用宽松匹配提取末尾可能的商品描述
                    pm = re.search(r'\s{2,}([\u4e00-\u9fa5a-zA-Z0-9()（）\-·【】\[\]]{2,50})\s*$', l)
                    if pm:
                        order.setdefault('托寄物内容1', pm.group(1))
                break
            if _looks_like_product(l):
                order.setdefault('托寄物内容1', l)
                break
            addr_parts.append(l)
        if addr_parts:
            # 对最后一部分也做一次商品剥离，同时尝试提取商品
            joined = ' '.join(addr_parts)
            last_addr = _strip_trailing_product(joined)
            order['收件详细地址'] = last_addr
            # 如果还没拿到商品，且剥离后有变化，再试一次
            if not order.get('托寄物内容1') and len(last_addr) < len(joined):
                pp = joined[len(last_addr):].strip()
                if pp:
                    order['托寄物内容1'] = pp
        return order

    return {}


# ========================================================================
# 格式D: "地址 姓名 电话"
# ========================================================================

def _extract_name_from_end(text, exclude_kw=None):
    """从文本末尾提取姓名，优先匹配称谓模式（X小姐/先生/女士）

    策略：
    1. 先在文本末尾查找称谓词（小姐/先生/女士等），找到后往前取1-2字作为姓
    2. 若无称谓，则用通用姓名模式（优先长匹配）

    Returns: (name, start_pos) 或 (None, -1)
    """
    if exclude_kw is None:
        exclude_kw = _get_address_exclude_kw()

    # 策略1: 称谓模式 — 先定位称谓词，再往前提取姓名（避免正则重叠问题）
    titles = ['小姐', '先生', '女士', '男士', '总经理', '经理', '老板',
              '帅哥', '美女', '阿姨', '叔叔', '婶婶', '大爷', '奶奶']
    for title in titles:
        if text.endswith(title):
            before_title = text[:-len(title)].rstrip()
            if before_title:
                # 中文姓氏绝大多数是单字（复姓如"欧阳"罕见），取末尾1字作为姓
                # 这样避免吃到地址词（如"...路石小姐"中的"路"）
                surname = before_title[-1]
                full_name = surname + title
                name_pos = len(before_title) - 1
                return full_name, name_pos
            break

    # 策略2: 通用姓名 — 长匹配优先（3>2>4），避免短字符串误匹配
    for length in [3, 2, 4]:
        m = re.search(rf'([\u4e00-\u9fa5]{{{length}}})\s*$', text)
        if m:
            cand = m.group(1)
            if not any(kw in cand for kw in exclude_kw):
                return cand, m.start()

    return None, -1


def _parse_addr_name_phone(block):
    """处理"地址 姓名 电话"格式"""
    lines = [l.strip() for l in block.split('\n') if l.strip()]
    if not lines:
        return {}

    order = {}
    first_line = lines[0]
    exclude_kw = _get_address_exclude_kw()

    # D1: 同行 — 从末尾找电话→逆向找姓名→剩余=地址
    phone_match = re.search(r'(1[3-9]\d{9})\s*$', first_line)
    if phone_match:
        before_phone = first_line[:phone_match.start()].rstrip()
        name, name_pos = _extract_name_from_end(before_phone, exclude_kw)
        if name:
            order['收件人'] = name
            order['收件电话'] = phone_match.group(1)
            addr_part = before_phone[:name_pos].strip()
            order['收件详细地址'] = _strip_trailing_product(addr_part)
            return order

    # D2: 跨行 "地址\n 姓名电话" 或 "地址\n 姓名(备注)\n 电话"
    if len(lines) >= 2:
        first_has_addr = bool(re.search(r'.{5,}(?:省|自治区|市).{0,15}(?:市|区|县|镇|乡|街|路)', lines[0]))

        # D2a: 标准两行 — 第二行有电话
        second_has_phone = bool(re.search(r'1[3-9]\d{9}', lines[1]))
        if first_has_addr and second_has_phone:
            name_m = re.match(r'^([\u4e00-\u9fa5]{2,4})', lines[1])
            phone_m = re.search(r'(1[3-9]\d{9})', lines[1])
            if name_m:
                order['收件人'] = name_m.group(1)
            if phone_m:
                order['收件电话'] = phone_m.group(1)
            addr_text = _clean_addr_line(lines[0], order.get('收件人'), order.get('收件电话'))
            order['收件详细地址'] = _strip_trailing_product(addr_text)
            return order

        # D2b: 三行 — 地址 + 姓名备注行(无电话) + 电话独占一行
        if first_has_addr and len(lines) >= 3:
            # 在后续行中找电话（跳过可能的姓名/备注行）
            phone_line_idx = None
            for idx in range(1, len(lines)):
                if re.search(r'^1[3-9]\d{9}$', lines[idx].strip()):
                    phone_line_idx = idx
                    break
            if phone_line_idx:
                # 从电话行之前的行提取姓名（可能是 "小陳:/有机素食工作室" 格式）
                name_from_prev = None
                for prev_idx in range(1, phone_line_idx):
                    prev_line = lines[prev_idx].strip()
                    # 支持 "名字:/备注" 或纯 "名字" 格式
                    name_m = re.match(r'^([\u4e00-\u9fa5]{2,4})(?::/|$|\s)', prev_line)
                    if name_m:
                        name_from_prev = name_m.group(1)
                        break
                    # 也尝试直接取前2-3个字作为姓名（如果这行很短且不含地址词）
                    if len(prev_line) <= 10 and re.match(r'^[\u4e00-\u9fa5/:：]+$', prev_line):
                        name_m2 = re.match(r'^([\u4e00-\u9fa5]{2,4})', prev_line)
                        if name_m2:
                            name_from_prev = name_m2.group(1)
                            break

                phone_m = re.search(r'(1[3-9]\d{9})', lines[phone_line_idx])
                if phone_m:
                    order['收件电话'] = phone_m.group(1)
                if name_from_prev:
                    order['收件人'] = name_from_prev
                addr_text = _clean_addr_line(lines[0], order.get('收件人'), order.get('收件电话'))
                order['收件详细地址'] = _strip_trailing_product(addr_text)
                return order

    # D3: 粘连 "长地址...姓名电话(无分隔)"
    stuck = re.search(r'(.{10,})([\u4e00-\u9fa5]{2,4})(1[3-9]\d{9})\s*$', first_line)
    if stuck:
        addr_part = stuck.group(1).strip()
        cand_name_raw = stuck.group(2)
        # 用称谓感知逻辑重新从addr_part+name中提取
        full_before_phone = addr_part + cand_name_raw
        name, name_pos = _extract_name_from_end(full_before_phone, exclude_kw)
        if name and name_pos > len(addr_part) - 2:
            # 姓名确实在地址之后
            order['收件人'] = name
            order['收件电话'] = stuck.group(3)
            order['收件详细地址'] = _strip_trailing_product(full_before_phone[:name_pos].strip())
            return order
        # 兜底：原始逻辑但用排除词检查
        if not any(kw in cand_name_raw for kw in exclude_kw):
            order['收件人'] = cand_name_raw
            order['收件电话'] = stuck.group(3)
            order['收件详细地址'] = _strip_trailing_product(addr_part)
            return order

    return {}



# ========================================================================
# 格式C2: 电话+姓名粘连（首行 = 电话+姓名，无空格）
# ========================================================================

def _parse_phone_name(block):
    """解析「电话+姓名粘连 地址 商品」格式

    典型输入：
      "13822153541廖琼"
      "广东省广州市南沙区南沙街道时代南湾7栋一单元2402"
      "羽衣甘蓝2.8斤"

    规则：
    - 首行：11位电话 + 中文姓名（粘连，无空格）
    - 第二行起：地址（含省市区街道等关键词）
    - 最后一行（可选）：商品+数量
    """
    order = {}
    lines = [l.strip() for l in block.split("\n") if l.strip()]
    if not lines:
        return {}

    # 1. 首行：电话+姓名粘连
    first = lines[0]
    m = re.match(r"^(1[3-9]\d{9})([\u4e00-\u9fa5]{2,4})$", first)
    if not m:
        return {}
    order["收件电话"] = m.group(1)
    order["收件人"] = m.group(2)

    # 2. 找地址行（含省市区镇街道路号栋单元等关键词）
    addr_lines = []
    remain_lines = []
    found_addr = False
    addr_kw = r"(?:省|自治区|市|区|县|镇|街道|路|巷|号|栋|单元|楼层|室|房|村|庄|苑|园|广场|大厦)"

    for line in lines[1:]:
        if not found_addr and re.search(addr_kw, line):
            addr_lines.append(line)
            found_addr = True
        elif found_addr:
            # 地址行之后，判断是否含数量单位（商品行特征）
            if re.search(r"[\d.]+\s*(?:斤|kg|g|个|袋|箱|份)", line):
                remain_lines.append(line)
                break
            else:
                addr_lines.append(line)
        else:
            remain_lines.append(line)

    if addr_lines:
        order["收件详细地址"] = "".join(addr_lines).strip()

    # 3. 剩余行提取托寄物内容（商品+规格整体填入，不剥离数量）
    product_text = " ".join(remain_lines).strip() if remain_lines else (lines[-1] if len(lines) > 2 else "")
    if product_text and not order.get("托寄物内容1"):
        order["托寄物内容1"] = product_text.strip()
        # 注：数量字段不自动填，规格已包含在托寄物内容1中

    return order

# ========================================================================
# 格式F: 联系电话前缀 (新增)
# ========================================================================

def _parse_contact_phone(block):
    """解析「地址 + 联系电话：phone + 姓名 + 商品」格式

    典型输入：
      "广东省 东莞市 寮步镇泰和路9号寮盈慧谷 4B栋402室   联系电话：18929132348   李生 蔬菜包6斤"
      "广州市天河区xx路xx号  联系电话： 13800138000  张三 有机玉米5斤"

    拆分策略：按「联系电话：」等前缀切分为地址段和剩余段，再从剩余段提取电话+姓名+商品。
    """
    order = {}
    exclude_kw = _get_address_exclude_kw()

    # 1. 找到联系电话前缀的位置
    contact_m = re.search(
        r'\s{2,}(?:联系电话|联系方式|收件电话|收货电话)\s*[：:]\s*',
        block
    )
    if not contact_m:
        return {}

    addr_part = block[:contact_m.start()].strip()
    after_label = block[contact_m.end():].strip()

    # 2. 从 after_label 中提取电话（开头的连续数字）
    phone_m = re.match(r'(\d{11})', after_label)
    if not phone_m:
        # 电话可能不在最开头，用宽松匹配
        phone_m2 = re.search(r'(1[3-9]\d{9})', after_label)
        if phone_m2:
            order['收件电话'] = phone_m2.group(1)
            after_phone_start = after_label[:phone_m2.start()].strip()
            after_phone_end = after_label[phone_m2.end():].strip()
            # 电话前的内容可能是空或额外信息
            remaining = (after_phone_start + ' ' + after_phone_end).strip()
        else:
            remaining = after_label
    else:
        order['收件电话'] = phone_m.group(1)
        remaining = after_label[phone_m.end():].strip()

    # 3. 设置地址
    if addr_part:
        order['收件详细地址'] = _strip_trailing_product(addr_part)

    # 4. 从 remaining 中提取姓名和商品
    if remaining:
        # 优先尝试 "姓名 商品" 模式（联系电话格式中姓名在前）
        m = re.match(r'^([\u4e00-\u9fa5]{2,4})\s+(.{2,50})$', remaining)
        if m:
            cand_name = m.group(1)
            if not any(kw in cand_name for kw in exclude_kw):
                order['收件人'] = cand_name
                order['托寄物内容1'] = m.group(2).strip()
        if not order.get('收件人'):
            # 也尝试从末尾提取姓名（兜底）
            name, name_pos = _extract_name_from_end(remaining, exclude_kw)
            if name:
                order['收件人'] = name
                before_name = remaining[:name_pos].strip()
                if before_name:
                    order['托寄物内容1'] = before_name
        if not order.get('收件人'):
            # 最后手段：直接按空格拆分第一个短词为姓名
            parts = remaining.split(None, 1)
            if len(parts[0]) <= 4 and re.match(r'^[\u4e00-\u9fa5]+$', parts[0]):
                cand = parts[0]
                if not any(kw in cand for kw in exclude_kw):
                    order['收件人'] = cand
                    if len(parts) > 1:
                        order['托寄物内容1'] = parts[1].strip()

    return order


# ========================================================================
# 格式E: 兜底
# ========================================================================

def _parse_fallback(block):
    """兜底策略"""
    order = {}
    kw_patterns = {
        '收件人': r'(?:收件人|收货人|姓名|联系人)[：:\s]*([^\n，,。\d]{2,8})',
        '收件电话': r'(?:收件电话|收货电话|手机|电话|联系电话|联系方式|手机号码)[：:\s]*(1[3-9]\d{9}|0\d{2,3}[-\s]?\d{7,8})',
        '收件详细地址': r'(?:收件地址|收货地址|地址|详细地址|送货地址)[：:\s]*([^\n]{5,80})',
        '用户订单号': r'(?:订单号|单号|订单编号)[：:\s]*([A-Za-z0-9\-]{4,25})',
        '寄方备注': r'(?:备注|备注信息|特殊说明)[：:\s]*([^\n]{1,50})',
    }
    for field, pattern in kw_patterns.items():
        m = re.search(pattern, block)
        if m:
            val = m.group(1).strip()
            if field == '收件详细地址' and val:
                val = _clean_address_field(val, order)
            order[field] = val
    if not order.get('收件电话'):
        phones = re.findall(r'1[3-9]\d{9}', block)
        if phones:
            order['收件电话'] = phones[0]
    if not order.get('收件详细地址'):
        addr_match = re.search(
            r'[^\s，,。；;\d]{2,5}?(?:省|自治区|市)\s*[^\s，,。；;\d]{0,10}?(?:市|州|区|县)[^\n，,。；;]{5,60}',
            block
        )
        if addr_match:
            order['收件详细地址'] = _strip_trailing_product(addr_match.group())
    if not order.get('收件人'):
        _extract_name_smart(order, block)
    return order


# ========================================================================
# 公共工具函数
# ========================================================================

def _get_address_exclude_kw():
    """获取地址清洗排除关键词（用于判断一个词是不是地名而非姓名）"""
    return [
        '省','市','区','县','路','街','号','楼','镇','乡','村',
        '单元','室','层','栋','大厦','广场','花园','公寓','小区',
        '收件','寄件','地址','公司','单位','街道'
    ]


def _strip_trailing_product(addr_text):
    """从地址末尾剥离可能粘连的商品描述

    例如: "...建设三马路32号604    新鲜有机甜玉米（10斤装）"
         → "...建设三马路32号604"
    """
    if not addr_text:
        return addr_text

    # 匹配末尾的商品描述模式：空白 + （中文/字母/数字/括号组成的商品名）
    patterns = [
        # "    新鲜有机甜玉米（10斤装）"
        r'\s{2,}([\u4e00-\u9fa5a-zA-Z0-9()（）\-【】\[\]·]{2,30}(?:\(\d+[斤kgKG公斤克g箱袋盒件装]+\)|【.*?】|\(\d+\))?)\s*$',
        # " 蔬菜包8种加玉米"
        r'\s+([\u4e00-\u9fa5a-zA-Z]{2,20}(?:\d*[斤kgKG公斤克g箱袋盒件]|加\w+)\s*)$',
        # "(10斤装)" 结尾的
        r'\s*\(?\d*\.?\d*\s*(?:斤|kg|KG|公斤|克|g|件|箱|袋|盒)(?:装)?\)?\s*$',
        # 任何以常见农产品词结尾且长度>4的尾巴
        r'\s{2,}([\u4e00-\u9fa5]{2,})\s*(?:\d+[斤kgKG公斤克g])?\s*$',
    ]

    cleaned = addr_text
    for pat in patterns:
        m = re.search(pat, cleaned)
        if m:
            # 确保不会过度截取（保留至少10个字符的地址）
            remaining_len = cleaned[:m.start()].strip().__len__()
            if remaining_len >= 10:
                cleaned = cleaned[:m.start()].strip()
            break

    return cleaned


def _clean_addr_line(addr_line, known_name=None, known_phone=None):
    """清理地址行：去掉已知的姓名和电话"""
    cleaned = addr_line
    if known_phone and known_phone in cleaned:
        cleaned = cleaned.replace(known_phone, '').strip()
    if known_name and known_name in cleaned:
        cleaned = cleaned.replace(known_name, '').strip()
    # 去掉多余空格
    cleaned = re.sub(r'\s{2,}', ' ', cleaned)
    return cleaned


def _looks_like_product(text):
    """判断一行文字是否看起来像是商品描述"""
    text = text.strip()
    if not text or len(text) < 1:
        return False

    # 排除非商品的词
    not_goods = [
        '地址', '电话', '手机', '联系人', '收件', '寄件',
        '备注', '订单', '单号', '编号', '公司', '单位',
        '省', '市', '区', '县', '路', '街', '号', '楼',
        '寄件人', '收货人'
    ]
    if any(ng in text for ng in not_goods):
        return False

    # 商品模式（任一匹配即判定为商品）
    prod_patterns = [
        # 以数字+单位开头的: "4.5斤玉米"、"10斤装"
        r'^\d+(?:\.\d+)?\s*(?:斤|kg|KG|公斤|克|g|件|箱|袋|盒)',
        # 纯中文商品名（2字以上）: "玉米"、"红菜头"、"蔬菜包"
        r'^[\u4e00-\u9fa5]{2,20}$',
        # 中文商品名 + 小数/整数 + 单位: "玉米4.5斤"、"红菜头3斤"
        r'^[\u4e00-\u9fa5]{2,30}\d+(?:\.\d+)?\s*(?:斤|kg|KG|公斤|克|g|件|箱|袋|盒|装)\s*$',
        # 带规格的商品名: "有机蔬菜 5斤"、"玉米5斤"
        r'^[\u4e00-\u9fa5a-zA-Z]{1,30}\s*\d+(?:\.\d+)?\s*(?:斤|kg|KG|公斤|克|g|件|箱|袋|盒|装)$',
        # 末尾为数字（无单位）的商品名: "茄子辣椒双拼4.5"、"玉米4.5"
        r'^[\u4e00-\u9fa5]{2,30}\d+(?:\.\d+)?$',
        # 带括号规格: "新鲜有机甜玉米（10斤装）"
        r'^[\u4e00-\u9fa5a-zA-Z0-9()（）\-·【】\[\]]{2,50}[\(（]',
        # 包/套餐/组合类: "蔬菜包8种加玉米"
        r'[\u4e00-\u9fa5]{2,}(?:包|套餐|组合|篮|礼盒|系列)',
    ]

    for pat in prod_patterns:
        if re.match(pat, text):
            return True

    return False


def _clean_address_field(addr_val, order):
    """清理地址字段：去掉末尾粘连的姓名+电话"""
    cleaned = addr_val
    # 如果地址末尾有11位手机号，截掉它及前面的姓名
    phone_at_end = re.search(r'(1[3-9]\d{9})\s*$', cleaned)
    if phone_at_end:
        before_phone = cleaned[:phone_at_end.start()].rstrip(' ，,；;、')
        name_before_phone = re.search(r'([\u4e00-\u9fa5]{2,4})\s*$', before_phone)
        exclude_kw = ['省','市','区','县','路','街','号','楼','镇','乡','村',
                      '单元','室','层','栋','大厦','广场']
        if name_before_phone:
            cand = name_before_phone.group(1)
            if not any(kw in cand for kw in exclude_kw):
                order.setdefault('收件人', cand)
                cleaned = before_phone[:name_before_phone.start()].rstrip(' ，,；;、')
        else:
            cleaned = before_phone
        order.setdefault('收件电话', phone_at_end.group(1))
    return cleaned.strip()


def _extract_name_smart(order, block):
    """从block中提取收件人姓名（智能版，避免误识别备注/别名）"""
    phone = order.get('收件电话', '')

    exclude_kw = [
        '省','市','区','县','路','街','号','楼','镇','乡','村',
        '单元','室','层','栋','大厦','广场','花园',
        '收件','寄件','地址','商品','货物','手机','电话','公司',
        '单位','订单','备注','数量','工作室','有机','素食','农场',
        '种植','基地','合作社',
        '联系','收货','送货','详细','寄件人',
    ]

    if phone:
        before = block[:block.find(phone)] if phone in block else block
        names = re.findall(r'[\u4e00-\u9fa5]{2,4}', before[-30:])
        for name in reversed(names):
            if not any(kw in name for kw in exclude_kw):
                order['收件人'] = name
                return
    else:
        lines = block.strip().split('\n')
        for line in lines:
            line = line.strip()
            if re.search(r'(?:省|自治区|市).*?(?:区|县|市)', line):
                continue
            if re.match(r'^[^：:\s]{2,10}[：:\s]', line):
                continue
            if _looks_like_product(line):
                continue
            m = re.match(r'^([\u4e00-\u9fa5]{2,4})', line)
            if m:
                cand = m.group(1)
                if not any(kw in cand for kw in exclude_kw):
                    order['收件人'] = cand
                    return


# 保留旧名作为兼容别名
def _extract_name_from_block(order, block):
    """兼容性别名"""
    return _extract_name_smart(order, block)


def _extract_product_from_other_lines(block, order):
    """从block的非地址行提取商品信息

    支持格式：
    - 玉米4.5斤
    - 新鲜有机甜玉米（10斤装）
    - 蔬菜包8种加玉米
    - 4.5斤玉米
    - 玉米
    """
    if order.get('托寄物内容1'):
        return

    lines = block.split('\n')
    # 跳过包含地址/电话信息的行（那些已经被解析为地址了）
    skip_patterns = [
        r'^(?:地址|收件地址|收货地址|送货地址|详细地址)',
        r'(?:省|自治区|市).*?(?:区|县|市)',
        r'1[3-9]\d{9}',
        r'^(?:收件人|收货人|姓名|联系人|手机|手机号码|电话|电话号码)',
        r'^(?:所在地区|地区|省市区)',
        r'^[^：:\s]{2,10}[：:\s]',  # 键值对行
    ]

    for line in lines:
        line = line.strip()
        if not line or len(line) < 2:
            continue
        # 跳过明显是地址/联系信息的行
        if any(re.search(p, line) for p in skip_patterns):
            continue
        # 商品模式（宽松匹配）
        if _looks_like_product(line):
            order['托寄物内容1'] = line.strip()
            break


def apply_defaults(orders, defaults):
    """将默认规则应用到所有订单"""
    # 这些字段不允许从默认值填入，必须来自订单本身
    NO_DEFAULT_FIELDS = {'托寄物内容1'}
    result = []
    for order in orders:
        merged = copy.deepcopy(defaults)
        # 移除不允许走默认值的字段
        for f in NO_DEFAULT_FIELDS:
            merged.pop(f, None)
        # 客户数据优先（覆盖默认值）
        for k, v in order.items():
            if v:
                merged[k] = v
        # 寄件人始终使用默认值，收件公司始终使用用户填写值，互不干扰
        result.append(merged)
    return result


def generate_excel(orders, duplicates=None):
    """基于顺丰官方模版生成填充后的xlsx"""
    try:
        wb = load_workbook(SF_TEMPLATE_PATH)
    except Exception:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '导入模板'
        ws.append(SF_COLUMNS)
    
    ws = wb['导入模板']
    
    # 找到表头行（第1行）
    header_row = [cell.value for cell in ws[1]]
    
    # 构建列名到列索引的映射
    col_map = {}
    for i, h in enumerate(header_row, 1):
        if h:
            col_map[h] = i
    
    # 从第2行开始写数据（第1行是表头）
    start_row = 2
    
    for row_idx, order in enumerate(orders):
        excel_row = start_row + row_idx
        for field, value in order.items():
            if field in col_map and value:
                ws.cell(row=excel_row, column=col_map[field], value=str(value))
    
    # 重复订单行黄色背景标记
    if duplicates:
        from openpyxl.styles import PatternFill
        yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for dup_idx in duplicates:
            excel_row = start_row + dup_idx
            for col in range(1, len(SF_COLUMNS) + 1):
                ws.cell(row=excel_row, column=col).fill = yellow
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ======================== 登录/登出路由 ========================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        # 返回登录页 HTML
        html = '''
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>登录 - 顺丰快递模版录入系统</title>
        <style>
          * { box-sizing: border-box; margin:0; padding:0; }
          body { font-family: -apple-system, BlinkMacSystemFont, 'PingFang SC', sans-serif;
                 background: linear-gradient(135deg, #e5000e 0%, #ff4444 100%);
                 min-height: 100vh; display: flex; align-items: center; justify-content: center; }
          .login-box { background: white; border-radius: 16px; box-shadow: 0 20px 60px rgba(0,0,0,0.15);
                        padding: 48px 40px; width: 380px; text-align: center; }
          .logo { color: #e5000e; font-size: 48px; margin-bottom: 8px; }
          .title { font-size: 20px; font-weight: 700; color: #1a1a2e; margin-bottom: 4px; }
          .subtitle { font-size: 13px; color: #6b7280; margin-bottom: 32px; }
          .field { text-align: left; margin-bottom: 18px; }
          .field label { display: block; font-size: 13px; color: #374151; font-weight: 500; margin-bottom: 6px; }
          .field input { width: 100%; padding: 10px 14px; border: 1px solid #d1d5db; border-radius: 8px;
                         font-size: 14px; outline: none; transition: border 0.2s; }
          .field input:focus { border-color: #e5000e; box-shadow: 0 0 0 3px rgba(229,0,14,0.1); }
          .btn { width: 100%; padding: 12px; background: linear-gradient(135deg, #e5000e, #ff4444);
                        border: none; border-radius: 8px; color: white; font-size: 15px; font-weight: 600;
                        cursor: pointer; transition: opacity 0.2s; }
          .btn:hover { opacity: 0.9; }
          .error { background: #fef2f2; color: #dc2626; padding: 10px 14px; border-radius: 8px;
                    font-size: 13px; margin-bottom: 16px; display: none; }
          .footer { margin-top: 24px; font-size: 11px; color: #9ca3af; }
        </style>
        </head>
        <body>
        <div class="login-box">
          <div class="logo">📦</div>
          <div class="title">顺丰快递模版录入系统</div>
          <div class="subtitle">请登录后使用</div>
          <div class="error" id="err"></div>
          <form method="POST" onsubmit="return doLogin(event)">
            <div class="field">
              <label>账号</label>
              <input type="text" name="username" id="username" placeholder="请输入账号" required autofocus>
            </div>
            <div class="field">
              <label>密码</label>
              <input type="password" name="password" id="password" placeholder="请输入密码" required>
            </div>
            <button class="btn" type="submit">登 录</button>
          </form>
          <div class="footer">顺丰模版录入系统 v2.0</div>
        </div>
        <script>
        function doLogin(e) {
          e.preventDefault();
          const err = document.getElementById('err');
          err.style.display = 'none';
          fetch('/api/login', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({
              username: document.getElementById('username').value.trim(),
              password: document.getElementById('password').value
            })
          }).then(r=>r.json()).then(d=>{
            if(d.success) { window.location.href = '/'; }
            else { err.innerText = d.error||'登录失败'; err.style.display='block'; }
          }).catch(()=>{
            err.innerText = '网络错误，请重试'; err.style.display='block';
          });
          return false;
        }
        </script>
        </body>
        </html>
        '''
        return html
    # POST 登录逻辑在 /api/login 处理


@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, username, is_admin FROM users WHERE username=? AND password=?', (username, password))
    user = c.fetchone()
    conn.close()
    if user:
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['is_admin'] = bool(user['is_admin'])
        return jsonify({'success': True, 'is_admin': bool(user['is_admin'])})
    return jsonify({'success': False, 'error': '账号或密码错误'}), 401


@app.route('/api/me', methods=['GET'])
def api_me():
    if 'user_id' in session:
        return jsonify({
            'username': session.get('username', ''),
            'is_admin': session.get('is_admin', False)
        })
    return jsonify({'username': '', 'is_admin': False})


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


# ======================== 订单数据 API ========================

@app.route('/api/orders/save', methods=['POST'])
@login_required
def save_orders():
    """保存订单到数据库（status=pending，未导出）
    采用全量覆盖策略：先删除该用户所有 pending 订单，再插入最新数据，避免重复累积。
    """
    data = request.json
    orders = data.get('orders', [])
    batch_id = data.get('batch_id', datetime.now().strftime('%Y%m%d%H%M%S'))
    user_id = session['user_id']
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    c = conn.cursor()
    # 全量覆盖：先删后插，避免重复累积
    c.execute('DELETE FROM orders WHERE user_id=? AND status=?', (user_id, 'pending'))
    for o in orders:
        c.execute(
            'INSERT INTO orders (user_id, order_data, status, batch_id, created_at) VALUES (?, ?, ?, ?, ?)',
            (user_id, json.dumps(o, ensure_ascii=False), 'pending', batch_id, now)
        )
    conn.commit()
    count = len(orders)
    conn.close()
    return jsonify({'success': True, 'saved': count, 'batch_id': batch_id})


@app.route('/api/orders/load', methods=['GET'])
@login_required
def load_orders():
    """加载当前用户的未导出订单（status=pending）"""
    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    c.execute(
        'SELECT id, order_data, batch_id, created_at FROM orders WHERE user_id=? AND status=? ORDER BY id ASC',
        (user_id, 'pending')
    )
    rows = c.fetchall()
    orders = []
    for row in rows:
        item = json.loads(row['order_data'])
        item['_db_id'] = row['id']
        orders.append(item)
    conn.close()
    return jsonify({'success': True, 'orders': orders})


@app.route('/api/orders/clear', methods=['POST'])
@login_required
def clear_orders():
    """标记已导出的订单为 exported（不清空，保留记录）"""
    data = request.json
    batch_id = data.get('batch_id', '')
    user_id = session['user_id']
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    c = conn.cursor()
    if batch_id:
        c.execute(
            'UPDATE orders SET status=?, exported_at=? WHERE user_id=? AND batch_id=? AND status=?',
            ('exported', now, user_id, batch_id, 'pending')
        )
    else:
        c.execute(
            'UPDATE orders SET status=?, exported_at=? WHERE user_id=? AND status=?',
            ('exported', now, user_id, 'pending')
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/orders/delete', methods=['POST'])
@login_required
def delete_order():
    """删除单条未导出订单"""
    data = request.json
    db_id = data.get('id')
    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM orders WHERE id=? AND user_id=? AND status=?', (db_id, user_id, 'pending'))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ======================== 用户管理 API（仅管理员） ========================

@app.route('/api/users/list', methods=['GET'])
@admin_required
def list_users():
    """获取所有用户列表（仅管理员）"""
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, username, is_admin, created_at FROM users ORDER BY id ASC')
    users = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'users': users})


@app.route('/api/users/add', methods=['POST'])
@admin_required
def add_user():
    """添加新用户（仅管理员）"""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    
    if not username or not password:
        return jsonify({'success': False, 'error': '用户名和密码不能为空'}), 400
    if len(username) < 2 or len(username) > 32:
        return jsonify({'success': False, 'error': '用户名长度需在 2-32 字符之间'}), 400
    if len(password) < 4:
        return jsonify({'success': False, 'error': '密码长度不能少于 4 位'}), 400
    if not re.match(r'^[\w\u4e00-\u9fa5]+$', username):
        return jsonify({'success': False, 'error': '用户名仅支持中文、字母、数字和下划线'}), 400
    
    conn = get_db()
    c = conn.cursor()
    # 检查是否已存在
    c.execute('SELECT id FROM users WHERE username = ?', (username,))
    if c.fetchone():
        conn.close()
        return jsonify({'success': False, 'error': f'用户 "{username}" 已存在'}), 409
    
    c.execute(
        'INSERT INTO users (username, password, is_admin, created_at) VALUES (?, ?, 0, ?)',
        (username, password, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    )
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'success': True, 'id': new_id, 'username': username})


@app.route('/api/users/delete', methods=['POST'])
@admin_required
def delete_user_api():
    """删除用户（仅管理员，不能删除自己）"""
    data = request.json
    target_id = data.get('id')
    if not target_id:
        return jsonify({'success': False, 'error': '请指定要删除的用户 ID'}), 400
    
    # 不能删除自己
    if int(target_id) == session.get('user_id'):
        return jsonify({'success': False, 'error': '不能删除当前登录的管理员账号'}), 400
    
    # 不能删除其他管理员
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, username, is_admin FROM users WHERE id = ?', (target_id,))
    target = c.fetchone()
    if not target:
        conn.close()
        return jsonify({'success': False, 'error': '用户不存在'}), 404
    if target['is_admin']:
        conn.close()
        return jsonify({'success': False, 'error': f'不能删除管理员 "{target["username"]}"'}), 403
    
    # 删除该用户的所有订单
    c.execute('DELETE FROM orders WHERE user_id = ?', (target_id,))
    # 删除用户
    c.execute('DELETE FROM users WHERE id = ?', (target_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'deleted': target['username']})


# ======================== 地址簿 API ========================

@app.route('/api/address/groups', methods=['GET'])
@login_required
def list_address_groups():
    """获取当前用户的地址簿分组列表"""
    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    c.execute(
        'SELECT id, name, sort_order, created_at FROM address_groups WHERE user_id=? ORDER BY sort_order ASC, id ASC',
        (user_id,)
    )
    groups = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'groups': groups})


@app.route('/api/address/groups', methods=['POST'])
@login_required
def create_address_group():
    """创建地址簿分组"""
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': '分组名称不能为空'}), 400
    if len(name) > 20:
        return jsonify({'success': False, 'error': '分组名称不能超过20个字符'}), 400

    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    # 获取当前最大 sort_order
    c.execute('SELECT COALESCE(MAX(sort_order), -1) + 1 AS next_order FROM address_groups WHERE user_id=?', (user_id,))
    next_order = c.fetchone()['next_order']
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute(
        'INSERT INTO address_groups (user_id, name, sort_order, created_at) VALUES (?, ?, ?, ?)',
        (user_id, name, next_order, now)
    )
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'success': True, 'id': new_id, 'name': name})


@app.route('/api/address/groups/<int:group_id>', methods=['PUT'])
@login_required
def rename_address_group(group_id):
    """重命名地址簿分组"""
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': '分组名称不能为空'}), 400
    if len(name) > 20:
        return jsonify({'success': False, 'error': '分组名称不能超过20个字符'}), 400

    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE address_groups SET name=? WHERE id=? AND user_id=?', (name, group_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/address/groups/<int:group_id>', methods=['DELETE'])
@login_required
def delete_address_group(group_id):
    """删除地址簿分组（同时删除其下所有地址）"""
    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    # 先删除该分组下的所有地址
    c.execute('DELETE FROM address_book WHERE group_id=? AND user_id=?', (group_id, user_id))
    # 再删除分组
    c.execute('DELETE FROM address_groups WHERE id=? AND user_id=?', (group_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/address/list', methods=['GET'])
@login_required
def list_addresses():
    """获取指定分组的地址列表"""
    group_id = request.args.get('group_id', type=int)
    if not group_id:
        return jsonify({'success': False, 'error': '请指定 group_id'}), 400

    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    c.execute(
        'SELECT id, group_id, name, phone, address, product, quantity, created_at FROM address_book WHERE user_id=? AND group_id=? ORDER BY id ASC',
        (user_id, group_id)
    )
    addresses = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'addresses': addresses})


@app.route('/api/address/add', methods=['POST'])
@login_required
def add_address():
    """添加地址到指定分组"""
    data = request.json
    group_id = data.get('group_id')
    if not group_id:
        return jsonify({'success': False, 'error': '请指定分组'}), 400

    name = data.get('name', '').strip()
    phone = data.get('phone', '').strip()
    address = data.get('address', '').strip()
    product = data.get('product', '').strip()
    quantity = data.get('quantity', '').strip() or '1'

    if not name or not phone or not address:
        return jsonify({'success': False, 'error': '收件人、电话、地址不能为空'}), 400

    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute(
        'INSERT INTO address_book (user_id, group_id, name, phone, address, product, quantity, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
        (user_id, group_id, name, phone, address, product, quantity, now)
    )
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'success': True, 'id': new_id})


@app.route('/api/address/<int:addr_id>', methods=['PUT'])
@login_required
def update_address(addr_id):
    """更新地址信息"""
    data = request.json
    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()

    updates = {}
    for field in ['name', 'phone', 'address', 'product', 'quantity', 'group_id']:
        if field in data:
            updates[field] = data[field].strip() if isinstance(data[field], str) else data[field]

    if not updates:
        return jsonify({'success': False, 'error': '没有需要更新的字段'}), 400

    set_clause = ', '.join(f'{k}=?' for k in updates)
    values = list(updates.values()) + [addr_id, user_id]
    c.execute(f'UPDATE address_book SET {set_clause} WHERE id=? AND user_id=?', values)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/address/<int:addr_id>', methods=['DELETE'])
@login_required
def delete_address(addr_id):
    """删除单条地址"""
    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM address_book WHERE id=? AND user_id=?', (addr_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/')
@login_required
def index():
    return app.send_static_file('index.html')


@app.route('/api/parse', methods=['POST'])
@login_required
def parse():
    """解析订单文本"""
    data = request.json
    raw_text = data.get('text', '')
    
    if not raw_text.strip():
        return jsonify({'success': False, 'error': '请输入订单信息'})
    
    try:
        orders = parse_orders(raw_text)
        return jsonify({'success': True, 'orders': orders, 'count': len(orders)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/generate', methods=['POST'])
@login_required
def generate():
    """生成Excel文件"""
    data = request.json
    orders = data.get('orders', [])
    defaults = data.get('defaults', {})
    
    if not orders:
        return jsonify({'success': False, 'error': '没有订单数据'})
    
    try:
        final_orders = apply_defaults(orders, defaults)
        duplicates = data.get('duplicates', [])
        excel_buf = generate_excel(final_orders, duplicates)
        
        date_str = datetime.now().strftime('%Y-%m-%d')
        filename = f'顺丰{date_str}.xlsx'
        
        return send_file(
            excel_buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/business_types', methods=['GET'])
@login_required
def business_types():
    """返回业务类型列表"""
    try:
        wb = load_workbook(SF_TEMPLATE_PATH)
        ws = wb['业务类型']
        types = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0]:
                types.append(row[0])
        return jsonify({'types': types})
    except Exception:
        default_types = ['顺丰特快', '顺丰标快', '顺丰即日', '顺丰空配', '专线普运', '重货包裹', '标准零担']
        return jsonify({'types': default_types})


# ======================== 统计看板 API ========================

def _stats_file_path(date_str=None):
    """获取指定日期的统计文件路径，默认今天"""
    if date_str is None:
        date_str = datetime.now().strftime('%Y-%m-%d')
    return os.path.join(STATS_DIR, f'stats_{date_str}.json')

def _cleanup_old_stats():
    """清理超过30天的统计文件"""
    cutoff = datetime.now() - timedelta(days=30)
    pattern = os.path.join(STATS_DIR, 'stats_*.json')
    for fpath in glob.glob(pattern):
        try:
            basename = os.path.basename(fpath)
            date_str = basename.replace('stats_', '').replace('.json', '')
            file_date = datetime.strptime(date_str, '%Y-%m-%d')
            if file_date < cutoff:
                os.remove(fpath)
        except (ValueError, OSError):
            pass  # 文件名格式不对或删除失败，跳过

@app.route('/api/stats/save', methods=['POST'])
@login_required
def save_stats():
    """保存当天导出统计（前端已做累加合并，后端直接写入）"""
    try:
        data = request.json
        stats = data.get('stats', {})
        today = datetime.now().strftime('%Y-%m-%d')
        
        payload = {
            'date': today,
            'exported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'orders_count': stats.get('orders_count', 0),
            'stats': stats.get('stats', []),
            'totalQty': stats.get('totalQty', 0),
            'totalCount': stats.get('totalCount', 0),
        }
        
        with open(_stats_file_path(today), 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        
        # 自动清理 30 天前的数据
        _cleanup_old_stats()
        
        return jsonify({'success': True, 'date': today})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/stats/today', methods=['GET'])
@login_required
def get_today_stats():
    """获取今天的统计数据"""
    today = datetime.now().strftime('%Y-%m-%d')
    fpath = _stats_file_path(today)
    if not os.path.exists(fpath):
        return jsonify({'success': True, 'data': None, 'date': today})
    try:
        with open(fpath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify({'success': True, 'data': data, 'date': today})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/stats/history', methods=['GET'])
@login_required
def get_history_stats():
    """获取指定日期的历史统计 ?date=2026-05-20"""
    date_str = request.args.get('date', '')
    if not date_str:
        return jsonify({'success': False, 'error': '请提供 date 参数'}), 400
    fpath = _stats_file_path(date_str)
    if not os.path.exists(fpath):
        return jsonify({'success': True, 'data': None, 'date': date_str})
    try:
        with open(fpath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify({'success': True, 'data': data, 'date': date_str})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/stats/dates', methods=['GET'])
def list_stats_dates():
    """列出所有有统计数据的日期（最近30天）"""
    pattern = os.path.join(STATS_DIR, 'stats_*.json')
    dates = []
    for fpath in glob.glob(pattern):
        try:
            basename = os.path.basename(fpath)
            date_str = basename.replace('stats_', '').replace('.json', '')
            # 验证日期格式
            datetime.strptime(date_str, '%Y-%m-%d')
            file_mtime = os.path.getmtime(fpath)
            dates.append({
                'date': date_str,
                'exported_at': datetime.fromtimestamp(file_mtime).strftime('%Y-%m-%d %H:%M:%S')
            })
        except (ValueError, OSError):
            pass
    dates.sort(key=lambda x: x['date'], reverse=True)
    return jsonify({'success': True, 'dates': dates})


@app.route('/api/upload', methods=['POST'])
def upload_and_parse():
    """统一入口：同时处理文本输入 + 多个表格文件，合并返回订单列表"""
    all_orders = []
    from_text = 0
    from_files = 0

    # ---- 1. 解析文本部分 ----
    text = request.form.get('text', '').strip()
    if text:
        try:
            text_orders = parse_orders(text)
            all_orders.extend(text_orders)
            from_text = len(text_orders)
        except Exception as e:
            pass

    # ---- 2. 解析上传的文件（支持多个）----
    uploaded_files = request.files.getlist('files[]')
    for f in uploaded_files:
        if not f or not f.filename:
            continue
        fname = f.filename.lower()
        try:
            if fname.endswith('.csv'):
                file_orders = _parse_csv_file(f)
            elif fname.endswith(('.xlsx', '.xls')):
                file_orders = _parse_excel_file(f)
            else:
                continue
            all_orders.extend(file_orders)
            from_files += len(file_orders)
        except Exception as e:
            # 单个文件失败不阻断其他
            continue

    if not all_orders:
        return jsonify({'success': False, 'error': '未能从文本或文件中识别到有效订单'})

    return jsonify({
        'success': True,
        'orders': all_orders,
        'count': len(all_orders),
        'from_text': from_text,
        'from_files': from_files,
    })


def _parse_csv_file(file_obj):
    """解析CSV/TSV文件为订单列表"""
    orders = []
    stream = io.TextIOWrapper(file_obj.stream, encoding='utf-8-sig')
    reader = csv.reader(stream)

    rows = [r for r in reader]
    if not rows:
        return orders

    lines = [','.join(r) for r in rows]  # 转成逗号分隔文本
    return parse_table(lines, ',')


def _parse_excel_file(file_obj):
    """解析Excel文件(xlsx/xls)为订单列表

    支持：
    - .xlsx: openpyxl 引擎
    - .xls: xlrd 引擎（旧版Excel格式，如德风导出）
    - 标准表格（表头在第1行）
    - 有赞/电商导出（表头在最后一行，如"收货人姓名"在底部）
    - 多Sheet
    """
    # === 根据文件扩展名选择解析引擎 ===
    fname = (file_obj.filename or '').lower()
    use_xlrd = fname.endswith('.xls') and not fname.endswith('.xlsx')

    suffix = '.xls' if use_xlrd else '.xlsx'
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_path = tmp.name
        file_obj.save(tmp_path)

    try:
        if use_xlrd:
            # 用 xlrd 读取旧版 .xls 格式
            import xlrd as _xlrd
            wb_xlrd = _xlrd.open_workbook(tmp_path)
            sheet_names = wb_xlrd.sheet_names()

            def _xlrd_cell_value(cell):
                """将 xlrd 单元格转为统一格式（类比 openpyxl values_only）"""
                if cell.ctype == _xlrd.XL_CELL_EMPTY:
                    return None
                if cell.ctype == _xlrd.XL_CELL_NUMBER:
                    val = cell.value
                    # 整数不保留 .0 后缀
                    return int(val) if val == int(val) else val
                if cell.ctype == _xlrd.XL_CELL_BOOLEAN:
                    return bool(cell.value)
                return cell.value

            _sheets_data = {}
            for sn in sheet_names:
                sh = wb_xlrd.sheet_by_name(sn)
                rows = []
                for r in range(sh.nrows):
                    row = tuple(_xlrd_cell_value(sh.cell(r, c)) for c in range(sh.ncols))
                    rows.append(row)
                _sheets_data[sn] = rows
        else:
            wb = load_workbook(tmp_path)
    except Exception as e:
        os.unlink(tmp_path)
        raise e

    # 表头关键词（用于检测哪一行是表头）
    header_keywords = [
        '收货人姓名', '收件人', '收货人', '联系人', '姓名',
        '联系电话', '收件电话', '收货电话', '手机', '电话',
        '收货地址', '收件地址', '地址', '详细地址',
        '商品名称', '商品规格', '商品', '货物',
        '订单号', '单号',
    ]

    # 非数据Sheet名称黑名单（说明页、备注页等不包含订单数据）
    _skip_sheet_patterns = {'说明', '填写说明', '备注', 'remark', 'note',
                            '帮助', 'help', '简介', '介绍', '导出说明',
                            '导出信息', 'export'}

    def _looks_like_header(row_cells):
        """判断一行是否看起来像表头（包含足够多的表头关键词）"""
        non_empty = [str(c).strip() for c in row_cells if c is not None and str(c).strip()]
        if not non_empty:
            return False, 0
        match_count = sum(
            1 for cell in non_empty
            for kw in header_keywords
            if kw in cell or cell in kw
        )
        # 至少匹配2个关键词才认为是表头行
        return match_count >= 2, match_count

    def _looks_like_data_row(row_cells):
        """判断一行是否像真实的订单数据行（而非说明/标题/空行）
        
        垃圾行特征：
        - 全行所有单元格都是短文本描述（如"说明"、"列表：每一行"）
        - 没有手机号、没有像样的姓名(2-4字汉字)、没有长地址文本
        """
        cells = [str(c).strip() for c in row_cells if c is not None and str(c).strip()]
        if not cells:
            return False
        # 如果任一单元格包含手机号，肯定是数据行
        import re
        for cell in cells:
            if re.search(r'1[3-9]\d{9}', cell):
                return True
        # 检查是否有像样长的内容（真实订单行通常有地址等长文本）
        long_cells = [c for c in cells if len(c) > 15]
        if len(long_cells) >= 1:
            return True
        # 所有单元格都很短（≤15字符），大概率是标题/说明行
        # 典型垃圾："说明", "采购商品列表：每一行为...", "列表"
        return False

    def _process_sheet(data_rows):
        """处理单个Sheet的数据行，返回订单列表"""
        if not data_rows or not any(data_rows[0]):
            return []

        total_rows = len(data_rows)
        if total_rows < 2:
            return []

        # === 检测表头位置：顶部(第1行) vs 底部(最后1行) ===
        first_is_header, first_score = _looks_like_header(data_rows[0])
        last_is_header, last_score = _looks_like_header(data_rows[-1])

        if last_is_header and not first_is_header:
            # 底部表头模式（有赞等电商导出）
            header_row = data_rows[-1]
            data_lines = data_rows[:-1]
        elif first_is_header:
            # 标准表头在顶部
            header_row = data_rows[0]
            data_lines = data_rows[1:]
        elif last_is_header and first_is_header:
            # 两头都像表头，优先用底部的（更可能是有赞格式）
            header_row = data_rows[-1]
            data_lines = data_rows[:-1]
        else:
            # 都不像表头，当作无表头数据处理
            header_row = None
            data_lines = data_rows

        # === 过滤掉非订单数据行（说明文字、标题行等）===
        if header_row is not None:
            data_lines = [row for row in data_lines if _looks_like_data_row(row)]
        else:
            data_lines = [row for row in data_rows if _looks_like_data_row(row)]

        lines = []
        if header_row is not None:
            cells = [str(c).strip() if c is not None else '' for c in header_row]
            lines.append('\t'.join(cells))

        for row_data in data_lines:
            cells = [str(c).strip() if c is not None else '' for c in row_data]
            lines.append('\t'.join(cells))

        return parse_table(lines, '\t')

    all_orders = []
    if use_xlrd:
        for sheet_name in sheet_names:
            sheet_clean = sheet_name.strip()
            if sheet_clean in _skip_sheet_patterns or any(p in sheet_clean for p in _skip_sheet_patterns):
                continue
            sheet_orders = _process_sheet(_sheets_data[sheet_name])
            all_orders.extend(sheet_orders)
    else:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_clean = sheet_name.strip()
            if sheet_clean in _skip_sheet_patterns or any(p in sheet_clean for p in _skip_sheet_patterns):
                continue
            data_rows = list(ws.iter_rows(values_only=True))
            sheet_orders = _process_sheet(data_rows)
            all_orders.extend(sheet_orders)

    os.unlink(tmp_path)
    return all_orders

# ======================== SF 导出订单分析 ========================

def _read_sf_export(file_stream):
    """读取顺丰导出订单 .xls/.xlsx 文件，提取运单号、收方联系人、收方公司

    参数: file_stream — Flask request.files 对象（file-like，有 save() 方法）
    返回: {公司名: [{'tracking': 'SF...', 'contact': '姓名'}, ...], ...}
    """
    tmp = tempfile.NamedTemporaryFile(suffix='.xls', delete=False)
    tmp_path = tmp.name
    tmp.close()
    file_stream.save(tmp_path)

    try:
        wb = xlrd.open_workbook(tmp_path)
        sh = wb.sheet_by_index(0)
    except Exception as e:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise ValueError(f'无法读取表格文件：{e}')

    COL_TRACKING = 3
    COL_CONTACT  = 19
    COL_COMPANY  = 22

    groups = {}
    for r in range(1, sh.nrows):
        tracking = str(sh.cell_value(r, COL_TRACKING)).strip()
        contact  = str(sh.cell_value(r, COL_CONTACT)).strip()
        company  = str(sh.cell_value(r, COL_COMPANY)).strip()

        if not tracking:
            continue

        company_key = company if company else '未标注公司'
        if company_key not in groups:
            groups[company_key] = []
        groups[company_key].append({
            'tracking': tracking,
            'contact': contact if contact else '(无姓名)',
        })

    if os.path.exists(tmp_path):
        os.unlink(tmp_path)

    return groups


@app.route('/sf-export')
def sf_export_page():
    """SF 导出订单分析页面"""
    return app.send_static_file('sf-export.html')


@app.route('/api/sf-export/parse', methods=['POST'])
def sf_export_parse():
    """上传 SF 导出文件，解析后按公司分组返回"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传文件'}), 400

    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'success': False, 'error': '文件为空'}), 400

    fname = f.filename.lower()
    if not (fname.endswith('.xls') or fname.endswith('.xlsx')):
        return jsonify({'success': False, 'error': '仅支持 .xls 或 .xlsx 格式'}), 400

    try:
        groups = _read_sf_export(f)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

    if not groups:
        return jsonify({'success': False, 'error': '未能从文件中提取到有效订单'}), 400

    result = []
    total = 0
    for company, orders in groups.items():
        result.append({
            'company': company,
            'count': len(orders),
            'orders': orders,
        })
        total += len(orders)

    result.sort(key=lambda x: -x['count'])

    return jsonify({
        'success': True,
        'total': total,
        'companies': len(result),
        'groups': result,
    })


@app.route('/api/sf-export/fill-tracking', methods=['POST'])
def sf_export_fill_tracking():
    """运单号回填：上传A表(顺丰导出)和多个B表(客户订单)，按姓名匹配，返回JSON数组（每项含base64 xlsx + 统计）

    请求: multipart/form-data
      - sf_file:      顺丰导出 xls（A表），含运单号+收方联系人
      - order_files:  客户订单 xlsx/xls（B表，可多个），含收件人姓名（任意格式）
    """
    import base64 as _b64

    sf_file     = request.files.get('sf_file')
    order_files = request.files.getlist('order_files')

    if not sf_file or not sf_file.filename:
        return jsonify({'success': False, 'error': '请上传顺丰导出文件（A表）'}), 400
    if not order_files:
        return jsonify({'success': False, 'error': '请至少上传一个客户订单文件（B表）'}), 400

    # ── Step 1: 从 A 表提取 {姓名: 运单号} 映射 ──────────────────────────────
    try:
        tmp_a = tempfile.NamedTemporaryFile(suffix='.xls', delete=False)
        tmp_a_path = tmp_a.name
        tmp_a.close()
        sf_file.save(tmp_a_path)

        wb_a = xlrd.open_workbook(tmp_a_path)
        sh_a = wb_a.sheet_by_index(0)
    except Exception as e:
        return jsonify({'success': False, 'error': f'A表读取失败：{e}'}), 400
    finally:
        if os.path.exists(tmp_a_path):
            os.unlink(tmp_a_path)

    COL_TRACKING = 3
    COL_CONTACT  = 19
    name_to_tracking = {}   # {姓名: 运单号}，同名取第一个
    for r in range(1, sh_a.nrows):
        tracking = str(sh_a.cell_value(r, COL_TRACKING)).strip()
        contact  = str(sh_a.cell_value(r, COL_CONTACT)).strip()
        if tracking and contact and contact not in name_to_tracking:
            name_to_tracking[contact] = tracking

    if not name_to_tracking:
        return jsonify({'success': False, 'error': 'A表中未找到有效的运单号/收件人信息'}), 400

    # 收件人别名列表（复用 parse_table 的 field_aliases）
    name_aliases = [
        '收货人/提货人姓名', '收件人姓名', '收货人姓名', '收件人', '收货人',
        '联系人', '姓名', '买家姓名', '客户姓名', 'name', 'receiver', '客户',
    ]

    def _find_name_col(header_line):
        """在表头行中找收件人姓名列，返回列索引或 None"""
        best_col = None
        best_score = 0
        best_len = 0
        for i, cell in enumerate(header_line):
            cell_c = str(cell).strip()
            if not cell_c:
                continue
            is_exact = any(a.lower() == cell_c.lower() for a in name_aliases)
            is_sub   = any(a.lower() in cell_c.lower() for a in name_aliases)
            is_in    = any(cell_c.lower() in a.lower() for a in name_aliases)
            if is_exact or is_sub or is_in:
                score = 3 if is_exact else (2 if is_sub else 1)
                hit_len = 0
                if is_exact:
                    hit_len = max(len(a) for a in name_aliases if a.lower() == cell_c.lower())
                elif is_sub:
                    hit_len = max(len(a) for a in name_aliases if a.lower() in cell_c.lower())
                else:
                    hit_len = len(cell_c)
                if score > best_score or (score == best_score and hit_len > best_len):
                    best_score = score
                    best_len = hit_len
                    best_col = i
        return best_col

    def _read_b_rows(order_file_obj):
        """读取B表为行列表，返回 (rows, error_msg)"""
        b_fname = order_file_obj.filename.lower()
        suffix_b = '.xls' if (b_fname.endswith('.xls') and not b_fname.endswith('.xlsx')) else '.xlsx'
        tmp_b = tempfile.NamedTemporaryFile(suffix=suffix_b, delete=False)
        tmp_b_path = tmp_b.name
        tmp_b.close()
        order_file_obj.save(tmp_b_path)
        try:
            if suffix_b == '.xls':
                import xlrd as _xlrd2
                wb_b_xlrd = _xlrd2.open_workbook(tmp_b_path)
                sh_b = wb_b_xlrd.sheet_by_index(0)
                rows = []
                for r in range(sh_b.nrows):
                    row = []
                    for c in range(sh_b.ncols):
                        cell = sh_b.cell(r, c)
                        val = cell.value
                        if cell.ctype == _xlrd2.XL_CELL_NUMBER and val == int(val):
                            val = int(val)
                        row.append(str(val).strip() if val != '' else '')
                    rows.append(row)
            else:
                wb_b = openpyxl.load_workbook(tmp_b_path)
                ws_b = wb_b.active
                rows = []
                for row in ws_b.iter_rows(values_only=True):
                    rows.append([str(c).strip() if c is not None else '' for c in row])
            return rows, None
        except Exception as e:
            return None, str(e)
        finally:
            if os.path.exists(tmp_b_path):
                os.unlink(tmp_b_path)

    # ── Step 2: 逐个处理每个 B 表 ─────────────────────────────────────────────
    results = []
    for order_file in order_files:
        orig_name = order_file.filename or 'unknown.xlsx'
        orig_stem = orig_name.rsplit('.', 1)[0] if '.' in orig_name else orig_name
        download_name = f'{orig_stem}_已填运单号_{datetime.now().strftime("%m%d")}.xlsx'

        b_rows, err = _read_b_rows(order_file)
        if err:
            results.append({'filename': orig_name, 'download_name': download_name,
                            'success': False, 'error': f'读取失败：{err}'})
            continue
        if len(b_rows) < 2:
            results.append({'filename': orig_name, 'download_name': download_name,
                            'success': False, 'error': 'B表数据不足（需至少1行表头+1行数据）'})
            continue

        name_col = _find_name_col(b_rows[0])
        if name_col is None:
            results.append({'filename': orig_name, 'download_name': download_name,
                            'success': False,
                            'error': '未找到收件人/姓名列，请确认表头包含"收件人"、"姓名"等字段'})
            continue

        # 插入运单号列
        matched = 0
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = 'Sheet1'
        for ri, row in enumerate(b_rows):
            if ri == 0:
                out_ws.append(['运单号'] + row)
            else:
                name = row[name_col].strip() if name_col < len(row) else ''
                tracking = name_to_tracking.get(name, '')
                if tracking:
                    matched += 1
                out_ws.append([tracking] + row)

        buf = io.BytesIO()
        out_wb.save(buf)
        xlsx_b64 = _b64.b64encode(buf.getvalue()).decode('utf-8')

        results.append({
            'filename': orig_name,
            'download_name': download_name,
            'success': True,
            'matched': matched,
            'total_b': len(b_rows) - 1,
            'xlsx_b64': xlsx_b64,
        })

    return jsonify({
        'success': True,
        'a_count': len(name_to_tracking),
        'results': results,
    })


@app.route('/api/sf-export/download', methods=['POST'])
def sf_export_download():
    """按公司下载 Excel — POST JSON: {company, orders}"""
    data = request.json or {}
    company = data.get('company', '').strip()
    orders  = data.get('orders', [])

    if not company:
        return jsonify({'success': False, 'error': '公司名称不能为空'}), 400
    if not orders:
        return jsonify({'success': False, 'error': '无订单数据'}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = company[:28]
    ws.append(['运单号', '收件人姓名'])

    for o in orders:
        ws.append([o.get('tracking', ''), o.get('contact', '')])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = company.replace('/', '／').replace('\\', '＼')[:50]
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{safe_name}_运单_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('DEBUG', 'true').lower() == 'true'
    app.run(debug=debug, port=port, host='0.0.0.0')
